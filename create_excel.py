import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import os
import re

# 读取问题文件
with open('questions_total.txt', 'r', encoding='utf-8') as f:
    lines = f.readlines()

# 解析问题
questions = []
for idx, line in enumerate(lines, start=1):
    line = line.strip()
    if line:
        # 直接使用行号作为编号，问题内容就是整行
        questions.append((idx, line))

print(f"找到 {len(questions)} 个问题")

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "问题与截图"

# 设置列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 80
ws.column_dimensions['C'].width = 60

# 设置表头
ws['A1'] = '编号'
ws['B1'] = '问题'
ws['C1'] = '截图'

# 设置表头样式
for cell in ['A1', 'B1', 'C1']:
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

print("开始处理问题和截图...")

# 设置行高（像素转换为磅）
row_height_pixels = 300
row_height_points = row_height_pixels * 0.75

# 遍历问题并添加到Excel
for idx, (num, question) in enumerate(questions, start=2):
    ws[f'A{idx}'] = num
    ws[f'B{idx}'] = question

    # 设置单元格对齐
    ws[f'A{idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 查找对应的截图文件
    screenshot_path = None
    screenshots_dir = 'screenshots'

    # 尝试匹配截图文件名
    for filename in os.listdir(screenshots_dir):
        if filename.startswith(f'question_{num}_'):
            screenshot_path = os.path.join(screenshots_dir, filename)
            break

    # 如果找到截图，插入到Excel
    if screenshot_path and os.path.exists(screenshot_path):
        try:
            img = Image(screenshot_path)

            # 调整图片大小以适应单元格
            max_width = 450
            max_height = 280

            # 计算缩放比例
            scale = min(max_width / img.width, max_height / img.height)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)

            # 设置行高
            ws.row_dimensions[idx].height = row_height_points

            # 插入图片到C列
            img.anchor = f'C{idx}'
            ws.add_image(img)

            if idx % 10 == 0:
                print(f"已处理 {idx-1} 个问题...")
        except Exception as e:
            print(f"处理问题 {num} 的截图时出错: {e}")
            ws[f'C{idx}'] = '图片加载失败'
    else:
        ws[f'C{idx}'] = '未找到截图'
        print(f"警告: 未找到问题 {num} 的截图")

print("保存Excel文件...")

# 保存Excel文件
output_file = '问题与截图汇总.xlsx'
wb.save(output_file)

print(f"完成！Excel文件已保存为: {output_file}")
print(f"共处理 {len(questions)} 个问题")
